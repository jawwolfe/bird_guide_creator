from guide_creater.utilites import SQLUtilities
from openpyxl import load_workbook
from guide_creater.exceptions import TaxonomyException
import shutil, datetime, csv, os


class GuideBase:
    def __init__(self, logger, sql_server_connection, guide_name, ebird_files, file_path, audio_path, image_path,
                 playlist_root):
        self.logger = logger
        self.sql_server_connection = sql_server_connection
        self.guide_name = guide_name
        self.ebird_files = ebird_files
        self.file_path = file_path
        self.audio_path = audio_path
        self.image_path = image_path
        self.playlist_root = playlist_root
        self.clements = None
        self.guide_id = None

    def get_clements(self):
        return self.clements

    def set_clements(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_clements')
        self.clements = utilities.run_sql_return_no_params()

    def set_guide_id(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_guide_id', params=' @GuideName=?', params_values=self.guide_name)
        a = utilities.run_sql_return_params()
        self.guide_id = utilities.run_sql_return_params()[0][0]

    def get_guide_id(self):
        return self.guide_id

    def _process_ebird_files(self):
        bird_list = []
        for file in self.ebird_files:
            my_wb = load_workbook(self.file_path + file)
            my_sheetname = "Sheet1"
            my_ws = my_wb[my_sheetname]
            raw_list = []
            for my_row in my_ws.iter_rows(min_row=2, values_only=True):
                raw_list.append(my_row[0])
            file_list = []
            for my_item in raw_list:
                if my_item:
                    # todo add another line to ignore the new page line characters
                    if 'sp.' not in my_item:
                        if '/' not in my_item:
                            if 'Domestic' not in my_item:
                                if 'undescribed' not in my_item:
                                    if my_item != 'Jan':
                                        if ' x ' not in my_item:
                                            file_list.append(my_item.replace('\t', ''))
            for item in file_list:
                bird_list.append(item)
        distinct_ebird_list = self._remove_ebird_duplicates(bird_list)
        return distinct_ebird_list

    def _remove_ebird_duplicates(self, mylist):
        distinct_ebird_data = []
        for bird in mylist:
            if bird not in distinct_ebird_data:
                distinct_ebird_data.append(bird)
        return distinct_ebird_data

    def _match_clements_ebird(self, clements, myebirds):
        # match to clements add scientific and taxon log any mismatches
        birds_clements = []
        for bird in myebirds:
            flag = False
            for taxon in clements:
                if taxon[1] == bird:
                    flag = True
                    diction = {'name': bird, 'code': taxon[4], 'scientific': taxon[2], 'target': ''}
                    birds_clements.append(diction)
            if not flag:
                msg = "This ebird bird English name" + bird + " didn't match the Clements taxonomy." \
                      " Please fix before proceeding"
                self.logger.error(msg)
                raise TaxonomyException(msg)
        return birds_clements

    def _check_new_add(self, mylist):
        flag = False
        for item in mylist:
            if item['add'] == 'ADD':
                flag = True
        return flag

    def _check_new_update(self, ebird, database):
        master_flag = False
        for bird in ebird:
            flag = False
            for name in database:
                if bird['name'] == name[1]:
                    flag = True
            if not flag:
                master_flag = True
        return master_flag

    def _create_playlists(self):
        playlists_sps = [{'sp': 'sp_get_in_guide', 'name': ''},
                         {'sp': 'sp_get_common_by_guide', 'name': 'Common Birds'},
                         {'sp': 'sp_get_common_uncommon_doves_by_guide', 'name': 'Common-Uncommon Doves and Cuckoos'},
                         {'sp': 'sp_get_common_passerines_by_guide', 'name': 'Common Passerines'},
                         {'sp': 'sp_get_common_scarce_passerines_by_guide', 'name': 'Common-Scarce Passerines'}]
        header = '#EXTM3U\n'
        item_begin = '#EXTINF:'
        extension = '.mp3\n'
        folder_phone = 'Birds/'
        # todo get phone root from database
        root_phone = '/storage/emulated/0/'
        playlist_path = self.playlist_root + self.guide_name
        if not os.path.exists(playlist_path):
            os.mkdir(playlist_path)
        for item in playlists_sps:
            if item['name'] == '':
                playlist_name = self.guide_name + ' Bird Guide'
            else:
                playlist_name = self.guide_name + ' ' + item['name']
            utilities = SQLUtilities(item['sp'], self.logger, sql_server_connection=self.sql_server_connection,
                                     params_values=self.guide_id, params='@GuideID=?')
            birds = utilities.run_sql_return_params()
            str_file = header
            for bird in birds:
                str_file += item_begin
                str_file += str(bird[2]) + ',' + bird[3] + " - "
                str_file += bird[0] + ' ' + bird[1] + '\n'
                str_file += root_phone + folder_phone + bird[0] + ' ' + bird[1] + extension
            f = open(playlist_path + "\\" + playlist_name + '.m3u', "w")
            f.write(str_file)


class CreateGuide(GuideBase):
    def __init__(self, ebird_files, file_path, logger, sql_server_connection, guide_name, audio_path, image_path,
                 playlist_root, exotic_file=None, targets_file=None):
        self.exotic_file = exotic_file
        self.targets_file = targets_file
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, guide_name=guide_name,
                           ebird_files=ebird_files, file_path=file_path, audio_path=audio_path, image_path=image_path,
                           playlist_root=playlist_root)

    def _process_exotic_file(self):
        return_list = []
        wb = load_workbook(self.file_path + self.exotic_file)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                flag = False
                code = None
                english = None
                for taxon in self.get_clements():
                    if taxon[2] == row[2]:
                        flag = True
                        code = taxon[4]
                        english = taxon[1]
                if flag:
                    diction = {'name': english, 'code': code, 'scientific': row[2], 'target': '', 'id': ''}
                    return_list.append(diction)
                else:
                    msg = "This exotic bird scientific name " + row[2] + " did not match Clements taxonomy, " \
                          "please fix before proceeding. English name: " + row[1]
                    self.logger.error(msg)
                    raise TaxonomyException(msg)
        return return_list

    def _get_targets(self):
        return_list = []
        wb = load_workbook(self.file_path + self.targets_file)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[1]:
                diction = {'name': row[1], 'scientific': row[2]}
                return_list.append(diction)
        return return_list

    def _combine_ebird_exotic(self, ebird_list, exotic_list):
        for exotic_bird in exotic_list:
            flag = False
            for ebird in ebird_list:
                if ebird['scientific'] == exotic_bird['scientific']:
                    flag = True
            if not flag:
                ebird_list.append(exotic_bird)
        return ebird_list

    def run_create(self):
        self.logger.info('Start script execution to create Bird Guide.')
        self.set_clements()
        self.set_guide_id()
        clements = self.get_clements()
        guide_id = self.get_guide_id()
        targets = None
        if self.targets_file:
            targets = self._get_targets()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        all_birds = utilities.run_sql_return_no_params()

        # collect all the ebird region files into one list then remove duplicates
        all_ebird_data = self._process_ebird_files()

        # match ebird list to clements on english name and get scientific name and taxon code
        # if using the same clements version year there should be no logged mismatches
        ebird_list_clements = self._match_clements_ebird(clements, all_ebird_data)

        # get exotic birds list and match to clements on scientific name, get clements english name and taxon code
        # log birds that don't match and fix these manually
        exotic_birds_clements = None
        if self.exotic_file:
            exotic_birds_clements = self._process_exotic_file()

        # if there are any exotic birds then add the diff to the ebird list (no duplicates)
        if exotic_birds_clements:
            all_birds_clements = self._combine_ebird_exotic(ebird_list_clements, exotic_birds_clements)
        else:
            all_birds_clements = ebird_list_clements

        # add targets
        if self.targets_file:
            for target in targets:
                for bird in all_birds_clements:
                    if 'Pheasant' in bird['name']:
                        pass
                    if target['scientific'] == bird['scientific']:
                        bird['target'] = 'TARGET'

        # compare the full list to birds already in guides database
        # and create a final list for adding to the database
        add_list = []
        for final_bird in all_birds_clements:
            flag = False
            id = None
            for master_bird in all_birds:
                if master_bird[1] == final_bird['name']:
                    flag = True
                    id = master_bird[0]
            if not flag:
                add = 'ADD'
            else:
                add = ''
            diction = {'name': final_bird['name'], 'code': final_bird['code'], 'scientific': final_bird['scientific'],
                       'add': add, 'target': final_bird['target'], 'id': id}
            add_list.append(diction)
        new_image_list = []
        # create paths for new images and audio files
        guide_des = 'New_Guide ' + self.guide_name + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        image_path = self.image_path
        audio_path = self.audio_path + guide_des + "\\"
        if self._check_new_add(add_list):
            os.mkdir(self.audio_path + guide_des)
        for add in add_list:
            target = 0
            if add['target'] == "TARGET":
                target = 1
            if add['add'] == "ADD":
                # add to birds database
                params_values = (add['name'], add['code'], add['scientific'], 1006)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird',
                                         params='@BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?')
                bird_id = utilities.run_sql_return_params()
                add['id'] = bird_id[0][0]
                self.logger.info("Added new bird to database: " + add['name'])
                params_values = (add['id'], guide_id, 1, 2, target, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                utilities.run_sql_params()
                self.logger.info("Added bird " + add['name'] + ' to the guide: ' + self.guide_name)
                diction = {'name': add['code'] + ' ' + add['name'], 'scientific': add['scientific']}
                new_image_list.append(diction)
                shutil.copy(self.audio_path + 'blank.mp3', audio_path + add['code'] + ' ' + add['name'] + '.mp3')

        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_in_guide', params='@GuideID=?', params_values=guide_id)
        guide_birds = utilities.run_sql_return_params()

        for add in add_list:
            flag = False
            target = 0
            if add['target'] == "TARGET":
                target = 1
            for name in guide_birds:
                if add['name'] == name[1]:
                    flag = True
            if not flag:
                params_values = (add['id'], guide_id, 1, 2, target, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                utilities.run_sql_params()
                self.logger.info("Added bird " + add['name'] + ' to the guide: ' + self.guide_name)

        if self._check_new_add(add_list):
            f = open(image_path + guide_des + '.csv', "w")
            for item in new_image_list:
                f.write('"' + item['name'] + '"' + ',"' + item['scientific'] + '"' + '\n')
            f.close()
            self._create_playlists()
            self.logger.info("Playlist updated.")

        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values=guide_id, params='@GuideID=?', sp='sp_update_guide_last_update')
        utilities.run_sql_params()
        self.logger.info('End Script Execution.\n')


class UpdateGuide(GuideBase):
    def __init__(self, ebird_files, file_path, logger, sql_server_connection, guide_name, audio_path, image_path,
                 playlist_root):
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection,
                           guide_name=guide_name, ebird_files=ebird_files, file_path=file_path, audio_path=audio_path,
                           image_path=image_path, playlist_root=playlist_root)

    def run_update(self):
        self.logger.info('Start script execution to update Bird Guide.')
        self.set_clements()
        self.set_guide_id()
        clements = self.get_clements()
        guide_id = self.get_guide_id()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        all_birds = utilities.run_sql_return_no_params()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_in_guide', params='@GuideID=?', params_values=guide_id)
        guide_birds = utilities.run_sql_return_params()

        # collect all the ebird region files into one list then remove duplicates
        all_ebird_data = self._process_ebird_files()

        # get clements data
        ebird_list_clements = self._match_clements_ebird(clements, all_ebird_data)

        # create paths for new images and audio files
        guide_des = 'Updates_Guide ' + self.guide_name + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        image_path = self.image_path
        audio_path = self.audio_path + guide_des + "\\"

        new_image_list = []
        # check for birds new to the birds database
        if self._check_new_update(ebird_list_clements, all_birds):
            os.mkdir(self.audio_path + guide_des)
        for ebird in ebird_list_clements:
            flag = False
            for name in all_birds:
                if ebird['name'] == name[1]:
                    ebird['id'] = name[0]
                    flag = True
            if not flag:
                print('Not in Birds Database:' + ebird['name'])
                # add to birds database
                params_values = (ebird['name'], ebird['code'], ebird['scientific'], 1006)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird',
                                         params='@BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?')
                bird_id = utilities.run_sql_return_params()
                ebird['id'] = bird_id[0][0]
                self.logger.info("Added new bird to database: " + ebird['name'])
                diction = {'name': ebird['code'] + ' ' + ebird['name'], 'scientific': ebird['scientific']}
                new_image_list.append(diction)
                shutil.copy(self.audio_path + 'blank.mp3', audio_path + ebird['code'] + ' ' + ebird['name'] + '.mp3')

        # check for birds new to the guide
        for ebird in ebird_list_clements:
            flag = False
            for name in guide_birds:
                if ebird['name'] == name[1]:
                    flag = True
            if not flag:
                print('Not in Birds Guide:' + ebird['name'])
                params_values = (ebird['id'], guide_id, 1, 2, 0, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                utilities.run_sql_params()
                self.logger.info("Added bird " + ebird['name'] + ' to the guide: ' + self.guide_name)

        if self._check_new_update(ebird_list_clements, all_birds):
            f = open(image_path + guide_des + '.csv', "w")
            for item in new_image_list:
                f.write('"' + item['name'] + '"' + ',"' + item['scientific'] + '"' + '\n')
            f.close()
            self._create_playlists()
            self.logger.info("Playlist updated.")

        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values=guide_id, params='@GuideID=?', sp='sp_update_guide_last_update')
        utilities.run_sql_params()
        self.logger.info("End script execution.")
