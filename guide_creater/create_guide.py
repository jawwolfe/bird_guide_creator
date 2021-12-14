from guide_creater.utilites import SQLUtilities
from openpyxl import load_workbook


class GuideBase:
    def __init__(self, logger, sql_server_connection):
        self.logger = logger
        self.sql_server_connection = sql_server_connection


class CreateGuide(GuideBase):
    def __init__(self, ebird_files, ebird_path, logger, sql_server_connection):
        self.ebird_files = ebird_files
        self.ebird_path = ebird_path
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection)

    def process_ebird_file(self, path):
        my_wb = load_workbook(path)
        my_sheetname = "Sheet1"
        my_ws = my_wb[my_sheetname]
        raw_list = []
        for my_row in my_ws.iter_rows(min_row=2, values_only=True):
            raw_list.append(my_row[0])
        return_list = []
        for my_item in raw_list:
            if my_item:
                # todo add another line to ignore the new page line characters
                if 'sp.' not in my_item:
                    if '/' not in my_item:
                        if 'Domestic' not in my_item:
                            if my_item != 'Jan':
                                if ' x ' not in my_item:
                                    return_list.append(my_item.replace('\t', ''))
        return return_list

    def remove_ebird_duplicates(self, mylist):
        distinct_ebird_data = []
        for bird in mylist:
            if bird not in distinct_ebird_data:
                distinct_ebird_data.append(bird)
        return distinct_ebird_data

    def match_clements(self, clements, mylist):
        # match to clements add scientific and taxon log any mismatches
        birds_clements = []
        for bird in mylist:
            flag = False
            for taxon in clements:
                if taxon[1] == bird:
                    flag = True
                    dict = {'name': bird, 'code': taxon[4], 'scientific': taxon[2]}
                    birds_clements.append(dict)
            if not flag:
                self.logger.inof('This bird did not match Clements: ' + bird)
        return birds_clements

    def run_guide(self):
        self.logger.info('Start script execution.')
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_clements')
        clements = utilities.run_sql_return_no_params()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        all_birds = utilities.run_sql_return_no_params()
        all_ebird_data = []
        for file in self.ebird_files:
            data = self.process_ebird_file(self.ebird_path + file)
            for item in data:
                all_ebird_data.append(item)
        distinct_ebird_list = self.remove_ebird_duplicates(all_ebird_data)
        dictinct_ebird_list_clements = self.match_clements(clements, distinct_ebird_list)
        add_list = []
        for ebird_bird in dictinct_ebird_list_clements:
            flag = False
            for master_bird in all_birds:
                if master_bird[1] == ebird_bird['name']:
                    flag = True
            if not flag:
                dict = {'name': ebird_bird['name'], 'code': ebird_bird['code'], 'scientific': ebird_bird['scientific']}
                add_list.append(dict)
        pass