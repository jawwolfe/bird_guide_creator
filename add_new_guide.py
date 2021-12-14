import pyodbc, csv, sys
from openpyxl import load_workbook
from exceptions import DatabaseConnectionException, DatabaseOperationException
from pyodbc import InterfaceError


def connect_sqlserver():
    connection_string = "Driver={ODBC Driver 17 for SQL Server};"
    connection_string += "Server=localhost\\SQLEXPRESS;"
    connection_string += "Database=BirdGuide;"
    connection_string += "Trusted_Connection=yes;"
    try:
        conn = pyodbc.connect(connection_string)
    except (pyodbc.OperationalError, InterfaceError) as err:
        msg = 'Error connecting to database via odbc'
        raise DatabaseConnectionException(msg)
    return conn


path_ebird = "C:\\temp\\ebird\\"

ebird_files = {'Ariana.xlsx', 'Ben Arous.xlsx', 'Bizerte.xlsx', 'Nabeul.xlsx', 'Tunis.xlsx'}


def process_ebird_file(path):
    my_wb = load_workbook(path)
    my_sheetname = "Sheet1"
    my_ws = my_wb[my_sheetname]
    raw_list = []
    for my_row in my_ws.iter_rows(min_row=2, values_only=True):
        raw_list.append(my_row[0])
    return_list = []
    for my_item in raw_list:
        if my_item:
            # todo add another line to ignore the new page line characters
            if 'sp.' not in my_item:
                if '/' not in my_item:
                    if 'Domestic' not in my_item:
                        if my_item != 'Jan':
                            if ' x ' not in my_item:
                                return_list.append(my_item.replace('\t', ''))
    return return_list


def get_clements_taxonomy():
    conn = connect_sqlserver()
    cursor = conn.cursor()
    sql = "Select * from Clements"
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
    except pyodbc.ProgrammingError as err:
        msg = "An error occurred executing a sql server command get birds"
        raise DatabaseOperationException(msg)
    conn.commit()
    return data


def get_current_birds():
    conn = connect_sqlserver()
    cursor = conn.cursor()
    sql = "Select * from birds;"
    try:
        cursor.execute(sql)
        data = cursor.fetchall()
    except pyodbc.ProgrammingError as err:
        msg = "An error occurred executing a sql server command get birds"
        raise DatabaseOperationException(msg)
    conn.commit()
    return data

c = 0
b = 0
a = 0

return_data = []

all_ebird_data = []
for file in ebird_files:
    data = process_ebird_file(path_ebird + file)
    for item in data:
        all_ebird_data.append(item)
distinct_ebird_data = []
for birds in all_ebird_data:
    if birds not in distinct_ebird_data:
        distinct_ebird_data.append(birds)

clements = get_clements_taxonomy()
for bird in distinct_ebird_data:
    flag = False
    taxon = None
    for taxon in clements:
        if taxon[1] == bird:
            flag = True
            taxon = taxon[4]
            print(bird)
            print(taxon)
    if not flag:
        a += 1
    else:
        b += 1
        taxon = taxon[4]
        dict = {'taxon': taxon, 'name': bird}
        return_data.append(dict)

    #sorted(return_data, key=lambda i: i['taxon'])


current_birds = get_current_birds()
for b in distinct_ebird_data:
    flag = False
    for cur in current_birds:
        if cur[1] == b:
            flag = True
    if not flag:
        print(b)
        new = 'ADD'
        c += 1
    else:
        pass

print(str(a))
print(str(b))
print(str(c))
