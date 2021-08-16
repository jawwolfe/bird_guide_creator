from openpyxl import load_workbook
import pyodbc

connection_string = "Driver={ODBC Driver 17 for SQL Server};"
connection_string += "Server=localhost\\SQLEXPRESS;"
connection_string += "Database=BirdGuide;"
connection_string += "Trusted_Connection=yes;"
conn = pyodbc.connect(connection_string)
path_taxonomy = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Clements_2019.xlsx'
path_new_island = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\samar_all.xlsx'


def get_scientific_name(bird_name, species):
    flag = False
    for species in species:
        if species['english'] == bird_name:
            flag = True
            return species['scientific']
    return ''


def get_bird_id(myname, mycode):
    sql = "select BirdID from Birds where BirdName = ? and TaxanomicCode = ?;"
    params = (myname, mycode)
    cursor = conn.cursor()
    cursor.execute(sql, params)
    bird_id = cursor.fetchone()
    if bird_id:
        return bird_id[0]
    else:
        return ''


def get_isalnd_id(myname):
    sql = "select IslandID from Islands where IslandeName = ?;"
    params = myname
    cursor = conn.cursor()
    cursor.execute(sql, params)
    island_id = cursor.fetchone()
    return island_id


def get_bird_island_id(bird_id, island_id):
    sql = "select ID from BirdsIslands where BirdID = ? and IslandID = ?;"
    params = (bird_id, island_id)
    cursor = conn.cursor()
    cursor.execute(sql, params)
    bird_island_id = cursor.fetchone()
    if bird_island_id:
        return bird_island_id[0]
    else:
        return ''


def get_closest_values(bird_id, island_id):
    sql = "EXEC sp_get_closest_difficulty_residence @birdid=?, @islandid=?"
    params = (bird_id, island_id)
    cursor = conn.cursor()
    cursor.execute(sql, params)
    values = cursor.fetchone()
    cursor.close()
    return values


def add_new_bird(myname, myprefix, myscientific):
    check_exists = get_bird_id(myname, myprefix)
    if not check_exists:
        sql = "Insert into Birds(BirdName, TaxanomicCode, ScientificName) values(?,?,?); Select @@Identity as 'ID';"
        params = (myname, myprefix, myscientific)
        cursor = conn.cursor()
        cursor.execute(sql, params)
        sql = "Select @@Identity as 'ID';"
        cursor.execute(sql)
        id = cursor.fetchone()[0]
        conn.commit()
        return id
    else:
        return check_exists


def add_bird_island(bird_id, island_id, mytarget, new):
    check_exists = get_bird_island_id(bird_id, island_id)
    residence = None
    difficulty = None
    if not new and not check_exists:
        values = get_closest_values(bird_id, island_id)
        residence = values[1]
        difficulty = values[0]
    if not check_exists:
        if not new:
            sql = 'Insert into BirdsIslands(BirdID, IslandID, ResidentStatusID, DifficultyID, IsTarget) values(?, ?, ?, ?, ?);'
            params = (bird_id, island_id, residence, difficulty, mytarget)
        else:
            sql = 'Insert into BirdsIslands(BirdID, IslandID, ResidentStatusID, DifficultyID, IsTarget) values(?, ?, 1, 3, ?);'
            params = (bird_id, island_id, mytarget)
        cursor = conn.cursor()
        cursor.execute(sql, params)
        conn.commit()


def update_bird_island_values(bird_id, island_id, myvalues):
    residence = myvalues[1]
    difficulty = myvalues[0]
    sql = 'Update BirdsIslands set DifficultyID = ?, ResidentStatusID = ? where BirdID = ? and IslandID = ?;'
    params = (difficulty, residence, bird_id, island_id)
    cursor = conn.cursor()
    cursor.execute(sql, params)
    conn.commit()


def update_bird_island_targets(bird_id, island_id, target):
    sql = 'Update BirdsIslands set IsTarget = ? where BirdID = ? and IslandID = ?;'
    params = (target, bird_id, island_id)
    cursor = conn.cursor()
    cursor.execute(sql, params)
    conn.commit()


# get the island from the file and go to database for ID
island_name = path_new_island[len('C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\'):-len('_all.xlsx')]
island = get_isalnd_id(island_name)[0]

wb = load_workbook(path_taxonomy)
sheetname = "eBird Clements v2019 Aug 2019"
ws = wb[sheetname]
clements_data = []
clements_species = []
for row in ws.iter_rows(min_row=2, values_only=True):
     data = {'category': row[2], 'english': row[3], 'scientific': row[4]}
     clements_data.append(data)
for item in clements_data:
    if item['category'] == "species":
        bird = {'english': item['english'], 'scientific': item['scientific']}
        clements_species.append(bird)

wb = load_workbook(path_new_island)
sheetname = "Sheet1"
ws = wb[sheetname]
new_island = []
for row in ws.iter_rows(min_row=1, values_only=True):
    data = {'code': row[0], 'english': row[1], 'scientific': row[2], 'add': row[3], 'target': row[4]}
    new_island.append(data)

'''
# For updating instead of inserting new as usual
for bird in new_island:
    prefix = bird['code']
    name = bird['english']
    target = bird['target']
    if target:
        target_value = 1
    else:
        target_value = 0
    myid = get_bird_id(name, prefix)
    exists = get_bird_island_id(myid, island)
    if exists:
        #update_bird_island_targets(myid, island, target_value)
        values = get_closest_values(myid, island)
        if values:
            update_bird_island_values(myid, island, values)
    else:
        print('NOT EXISTS')
'''


for bird in new_island:
    prefix = bird['code']
    name = bird['english']
    scientific = get_scientific_name(name, clements_species)
    if bird['target']:
        if bird['target'].strip() == 'TARGET':
            target_value = 1
        else:
            target_value = 0
    else:
        target_value = 0
    if not scientific:
        raise ValueError('No match on common name in Clements, check name for ' + name)
    if bird['add'] == 'ADD':
        myid = add_new_bird(name, prefix, scientific)
        add_bird_island(myid, island, target_value, new=True)
    else:
        myid = get_bird_id(name, prefix)
        if myid:
            add_bird_island(myid, island, target_value, new=False)
        else:
            raise ValueError("Cant find bird ID")
