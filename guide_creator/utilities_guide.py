from guide_creator.utilites import SQLUtilities
from openpyxl import load_workbook
import datetime, os, glob, shutil


class Compare:
    def __init__(self, logger, path_one, path_two=None, sql_server_connection=None):
        self.logger = logger
        self.path_one = path_one
        self.path_two = path_two
        self.sql_server_connection = sql_server_connection

    def process_ebird_file(self, path):
        wb = load_workbook(path)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        return_list = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            data = {'name': row[1], 'code': row[0]}
            return_list.append(data)
        return return_list

    def run_compare(self):
        lst_3 = self.process_ebird_file(self.path_one)
        lst_4 = self.process_ebird_file(self.path_two)
        add = []
        for item in lst_3:
            flag = False
            for item2 in lst_4:
                if item['name'] == item2['name']:
                    flag = True
            if not flag:
                diction = {'code': item['code'], 'name': item['name']}
                add.append(diction)

    def run_compare_db_directory(self):
        utilities = SQLUtilities(sp='sp_get_new_birds_by_guide', sql_server_connection=self.sql_server_connection,
                                 params_values=12, params='@GuideID=?', logger=self.logger)
        new_bird_list = utilities.run_sql_return_params()
        for item in new_bird_list:
            flag = False
            for filename in os.listdir(self.path_one):
                split = filename.split("_", 1)
                if item[0] == split[0]:
                    flag = True
            if not flag:
                print(item[0])


class RecreateImageList:
    def __init__(self, logger, sql_server_connection):
        self.logger = logger
        self.sql_server_connection = sql_server_connection

    def run_recreate_image_list(self):
        # this was to recreate the csv file with the new birds for a guide to create images
        utilities = SQLUtilities(sp='sp_get_new_birds_by_guide', sql_server_connection=self.sql_server_connection,
                                 params_values=1014, params='@GuideID=?', logger=self.logger)
        new_image_list = utilities.run_sql_return_params()

        guide_des = 'New_Guide ' + 'Kuala Lumpur and Selangor' + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        image_path = 'C:\\temp\\Source\\'

        f = open(image_path + guide_des + '.csv', "w")
        for item in new_image_list:
            f.write('"' + item[0] + '"' + ',"' + item[1] + '"' + ',"' + item[2] + '"' + '\n')
        f.close()


class RecreateAudioFiles:
    def __init__(self, logger, sql_server_connection):
        self.logger = logger
        self.sql_server_connection = sql_server_connection

    def run_recreate_audio_files(self):
        # this was to recreate the csv file with the new birds for a guide to create images
        utilities = SQLUtilities(sp='sp_get_new_birds_by_guide', sql_server_connection=self.sql_server_connection,
                                 params_values=1014, params='@GuideID=?', logger=self.logger)
        new_bird_list = utilities.run_sql_return_params()
        guide_des = 'New_Guide ' + 'Kuala Lumpur and Selangor' + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        audio_path = 'C:\\temp\\Audio\\'
        os.mkdir(audio_path + guide_des + '\\')
        for item in new_bird_list:
            shutil.copy(audio_path + 'blank.mp3', audio_path + guide_des + "\\" + item[0] + '.mp3')


class Rename:
    def __init__(self, logger, sql_server_connection, path_audio='', path_images=''):
        self.logger = logger
        self.sql_server_connection = sql_server_connection
        self.path_audio = path_audio
        self.path_images = path_images

    def run_rename(self):
        sql = "exec sp_get_names;"
        utilities = SQLUtilities(sp='sp_get_names', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection)
        names = utilities.run_sql_return_no_params()

        os.chdir(self.path_audio)
        for file in glob.glob('*'):
            old_path = self.path_audio + file
            if file[0] == 'z':
                new_path = self.path_audio + file[1:].replace('_', "'")
            else:
                new_path = self.path_audio + file.replace('_', "'")
            print(new_path)
            os.rename(old_path, new_path)
            '''
            print(file)
            for item in names:
                if item == item[0]:
                    old_path = self.path_audio + '.mp3'
                    new_path = self.path_audio + item[1] + '.mp3'
                    c += 1
                    print(new_path)
                    #os.rename(old_path, new_path)
            '''

        '''
        for item in data:
            if item[0] != item[1]:
                old_path_image = path_photos + item[2] + ' ' + item[1] + '.jpg'
                new_path_image = path_photos + item[2] + ' ' + item[0] + '.jpg'
                os.rename(old_path_image, new_path_image)
                old_path_audio = path_audio + item[2] + ' ' + item[1] + '.mp3'
                new_path_audio = path_audio + item[2] + ' ' + item[0] + '.mp3'
                os.rename(old_path_audio, new_path_audio)'''
