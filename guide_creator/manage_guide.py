from guide_creator.utilites import SQLUtilities
from openpyxl import load_workbook
from guide_creator.exceptions import TaxonomyException
import shutil, datetime, csv, os
from bs4 import BeautifulSoup
import requests, json, sys


class GuideBase:
    def __init__(self, logger, sql_server_connection, file_path):
        self.logger = logger
        self.sql_server_connection = sql_server_connection
        self.clements = None
        self.guide_id = None
        self.all_birds = None
        self.regions_birds = None
        self.exotic_guides_birds = None
        self.file_path = file_path

    def get_clements(self):
        return self.clements

    def set_clements(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_clements')
        self.clements = utilities.run_sql_return_no_params()

    def set_guide_id(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_guide_id', params=' @GuideName=?', params_values=self.guide_name)
        self.guide_id = utilities.run_sql_return_params()[0][0]

    def get_guide_id(self):
        return self.guide_id

    def set_all_birds(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        self.all_birds = utilities.run_sql_return_no_params()

    def get_all_birds(self):
        return self.all_birds

    def set_regions_birds(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_regions_birds')
        self.regions_birds = utilities.run_sql_return_no_params()

    def get_exotic_guides_birds(self):
        return self.exotic_guides_birds

    def set_exotic_guides_birds(self):
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                     sp='sp_get_exotic_guides_birds')
        self.exotic_guides_birds = utilities.run_sql_return_no_params()

    def get_regions_birds(self):
        return self.regions_birds

    def _process_ebird_files(self):
        bird_list = []
        utilities = SQLUtilities('sp_get_regions_by_guide_name', self.logger, self.sql_server_connection,
                                 params=' @GuideName=?', params_values=self.guide_name)
        regions = utilities.run_sql_return_params()
        for region in regions:
            file = region[1].replace(' ', '_') + '_' + region[0].replace(' ', '_') + '_' + 'Ebird.xlsx'
            my_wb = load_workbook(self.file_path + file)
            my_sheetname = "Sheet1"
            my_ws = my_wb[my_sheetname]
            raw_list = []
            for my_row in my_ws.iter_rows(min_row=1, values_only=True):
                raw_list.append(my_row[0])
            file_list = []
            for my_item in raw_list:
                if my_item:
                    # todo add another line to ignore the new page line characters
                    if 'sp.' not in my_item:
                        if '/' not in my_item:
                            if 'Domestic' not in my_item:
                                if 'undescribed' not in my_item:
                                    if my_item != 'Jan':
                                        if my_item != 'bird sp':
                                            if ' x ' not in my_item:
                                                file_list.append(my_item.replace('\t', ''))
            for item in file_list:
                bird_list.append(item)
        distinct_ebird_list = self._remove_ebird_duplicates(bird_list)
        return distinct_ebird_list

    def _remove_ebird_duplicates(self, mylist):
        distinct_ebird_data = []
        for bird in mylist:
            if bird not in distinct_ebird_data:
                distinct_ebird_data.append(bird)
        return distinct_ebird_data

    def _match_clements_ebird(self, clements, myebirds):
        # match to clements add scientific and taxon log any mismatches
        birds_clements = []
        for bird in myebirds:
            flag = False
            for taxon in clements:
                if taxon[1] == bird:
                    flag = True
                    diction = {'name': bird, 'code': taxon[4], 'scientific': taxon[2], 'target': ''}
                    birds_clements.append(diction)
            if not flag:
                msg = "This ebird bird English name" + bird + " didn't match the Clements taxonomy." \
                      " Please fix before proceeding"
                self.logger.error(msg)
                raise TaxonomyException(msg)
        return birds_clements

    def _check_new_add(self, mylist):
        flag = False
        for item in mylist:
            if item['add'] == 'ADD':
                flag = True
        return flag

    def _check_new_update(self, ebird, database):
        master_flag = False
        for bird in ebird:
            flag = False
            for name in database:
                if bird['name'] == name[1]:
                    flag = True
            if not flag:
                master_flag = True
        return master_flag


class CreateImageAudioTodoList(GuideBase):
    def __init__(self, logger, audio_guide_path, sql_server_connection, todo_path, file_path=None):
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, file_path=file_path)
        self.audio_guide_path = audio_guide_path
        self.todo_path = todo_path

    def run(self):
        self.logger.info("Begin script execution.")
        utilities = SQLUtilities(sp='sp_get_all_birds', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        birds_database = utilities.run_sql_return_no_params()
        f = open(self.todo_path + 'Images_Needed' + '.csv', "w")
        # delete all blank images before refreshing
        for audio in os.listdir(self.todo_path + 'Audio'):
            os.remove(self.todo_path + 'Audio\\' + audio)
        for bird in birds_database:
            flag = False
            str_rg = ''
            for file in os.listdir(self.audio_guide_path):
                if file.endswith(".mp3"):
                    prefix = file[0:4].strip()
                    name = file[5:][:-4]
                    if bird[1] == name and bird[2] == prefix:
                        flag = True
            if not flag:
                params = (bird[0])
                utilities = SQLUtilities(sp='sp_get_guides_regions_new', logger=self.logger,
                                         sql_server_connection=self.sql_server_connection,
                                         params_values=params, params=' @BirdID=?')
                regions_guides = utilities.run_sql_return_params()
                for rg in regions_guides:
                    str_rg += rg[0] + ', '
                f.write('"' + bird[2] + ' ' + bird[1] + '"' + ',"' + bird[3] + '"' + ',"' + str_rg + '"' + '\n')
                shutil.copy(self.todo_path + 'blank.mp3', self.todo_path + 'Audio\\' + bird[2] + ' ' + bird[1] + '.mp3')
        f.close()
        self.logger.info("End script execution.")


class ExoticParseUtility(GuideBase):
    def __init__(self, logger, exotic_base_url, sql_server_connection, file_path=None):
        self.exotic_base_url = exotic_base_url
        self.sql_server_connection = sql_server_connection
        self.file_path = file_path
        self.targets = None
        self.specialities = None
        self.errors = None
        self.pages = ['/checklist.html', '/target-birds.html', '/special-birds.html']
        self.exclude = ['Streptopelia bitorquata', 'Aerodramus vanikorensis']
        self.char_map = {'/': 2, '\\': 3, '|': 8, '#': 7, '<': 6, '(': 4, '{': 5, '[': 9}
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, file_path=file_path)

    def get_errors(self):
        return self.errors

    def set_errors(self):
        utilities = SQLUtilities(sp='sp_get_exotic_errors', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        errors = utilities.run_sql_return_no_params()
        self.errors = errors

    def get_targets(self, guide_id, scientific):
        return_data = []
        flag = False
        for item in self.targets:
            if item['scientific'] == scientific and item['guide'] == guide_id:
                flag = True
                return_data.append(item['likelihood'])
                return_data.append(1)
        if not flag:
            return_data.append('')
            return_data.append(0)
        return return_data

    def get_specialities(self, guide_id, scientific):
        return_data = []
        flag = False
        endemic_map = {'E': 1, 'NE': 2}
        conservation_map = {'NT': 2, 'V': 3, 'EN': 4, 'CR': 5}
        for item in self.specialities:
            if item['scientific'] == scientific and item['guide'] == guide_id:
                flag = True
                edem_val = ''
                cons_val = ''
                for key, val in endemic_map.items():
                    if key == item['endemic']:
                        edem_val = val
                for key, val in conservation_map.items():
                    if key == item['conservation']:
                        cons_val = val
                return_data.append(edem_val)
                return_data.append(cons_val)
        if not flag:
            return_data.append(5)
            return_data.append(1)
        return return_data

    def set_targets(self):
        all_data = []
        utilities = SQLUtilities(sp='sp_get_guides', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        guides = utilities.run_sql_return_no_params()
        for guide in guides:
            guide_id = guide[0]
            # Get the checklist for this guide from exotic website
            url = self.exotic_base_url + guide[2] + self.pages[1]
            website = requests.get(url)
            soup = BeautifulSoup(website.content, 'html5lib')
            my_tables = soup.findAll("table")
            rows = my_tables[1].findChildren(['tr'])
            for row in rows:
                tds = row.find_all('td')
                diction = {'guide': guide_id, 'scientific': str(tds[2].find('i').contents[0].strip()),
                           'likelihood': str(tds[3].contents[0].strip())}
                all_data.append(diction)
        self.targets = all_data

    def set_specialities(self):
        all_data = []
        utilities = SQLUtilities(sp='sp_get_guides', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        guides = utilities.run_sql_return_no_params()
        for guide in guides:
            guide_id = guide[0]
            # Get the checklist for this guide from exotic website
            url = self.exotic_base_url + guide[2] + self.pages[2]
            website = requests.get(url)
            soup = BeautifulSoup(website.content, 'html5lib')
            my_tables = soup.findAll("table")
            rows = my_tables[1].findChildren(['tr'])
            for row in rows:
                tds = row.find_all('td')
                diction = {'guide': guide_id, 'scientific': str(tds[2].find('i').contents[0].strip()),
                           'endemic': str(tds[3].contents[0].strip()), 'conservation': str(tds[4].contents[0].strip())}
                all_data.append(diction)
        self.specialities = all_data

    def parse_chars(self, first_char):
        return_value = ''
        flag = False
        for key, val in self.char_map.items():
            if first_char == key:
                return_value = val
                flag = True
        if not flag:
            return_value = 1
        return return_value

    def parse_all_guides(self):
        self.logger.info('Start Execution.')
        self.set_clements()
        self.set_all_birds()
        self.set_exotic_guides_birds()
        self.set_targets()
        self.set_specialities()
        utilities = SQLUtilities(sp='sp_get_guides', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        guides = utilities.run_sql_return_no_params()
        c = 0
        for guide in guides:
            guide_id = guide[0]
            self.set_errors()
            # Get the checklist for this guide from exotic website
            url = self.exotic_base_url + guide[2] + self.pages[0]
            website = requests.get(url)
            soup = BeautifulSoup(website.content, 'html5lib')
            my_tables = soup.findAll("table")
            rows = my_tables[1].findChildren(['tr'])
            row_ct = 0
            for row in rows:
                if row_ct > 0:
                    bird_id = None
                    code = ''
                    bird_name_clements = ''
                    tds = row.find_all('td')
                    if len(tds) > 1:
                        flag_clement_match = False
                        flag_birds_match = False
                        scientific_name = tds[2].find('i').contents[0].strip()
                        c += 1
                        bird_name_exotic_raw = tds[1].contents[0].strip()
                        if bird_name_exotic_raw[0] in ['/', '\\', '|', '#', '(', '{', '[', '<'] and bird_name_exotic_raw[:-1][-2:] == '**':
                            bird_name_exotic = bird_name_exotic_raw[1:-1]
                            bird_name_exotic = bird_name_exotic[:-2]
                        elif bird_name_exotic_raw[0] in ['/', '\\', '|', '#', '(', '{', '[', '<']:
                            bird_name_exotic = bird_name_exotic_raw[1:-1]
                        elif bird_name_exotic_raw[-2:] == '**':
                            bird_name_exotic = bird_name_exotic_raw[:-2]
                        else:
                            bird_name_exotic = bird_name_exotic_raw
                        for taxon in self.get_clements():
                            # need to match on both Scientific and common name
                            if taxon[2] == scientific_name.strip() and taxon[1] == bird_name_exotic:
                                if taxon[2] not in self.exclude:
                                    flag_clement_match = True
                                    code = taxon[4]
                                    bird_name_clements = taxon[1]
                        first_char = bird_name_exotic_raw[0]
                        res_status_id = self.parse_chars(first_char)
                        if not flag_clement_match:
                            self.logger.info('Exotic bird name not found in Clements: ' + bird_name_exotic +
                                             ' checking for duplicate error.')
                            # add this bird to error table and break out of loop and go to next bird
                            # check for duplicates
                            fl_error = False
                            for error in self.get_errors():
                                if error[1] == scientific_name and error[3] == guide_id:
                                    fl_error = True
                            if not fl_error:
                                params = (bird_name_exotic, guide_id, res_status_id, scientific_name)
                                utilities = SQLUtilities(logger=self.logger,
                                                         sql_server_connection=self.sql_server_connection,
                                                         sp='sp_insert_exotic_error',
                                                         params=' @BirdName=?,@GuideID=?,@ResidentStatusID=?, '
                                                                '@ScientificName=?',
                                                         params_values=params)
                                utilities.run_sql_params()
                            continue
                        # get the bird ID with scientific name
                        for bird in self.all_birds:
                            if bird[3] == scientific_name.strip():
                                flag_birds_match = True
                                bird_id = bird[0]
                        if not flag_birds_match:
                            # enter into birds table then exotic table
                            params = (bird_name_clements.strip(), code, scientific_name, 1006)
                            utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                     sp='sp_insert_bird',
                                                     params=' @BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?',
                                                     params_values=params)
                            bird_id = utilities.run_sql_return_params()[0][0]
                            self.logger.info("Added a new bird to the Birds table: " + bird_name_clements.strip() +
                                             ' ,id: ' + str(bird_id))
                            # refresh the all birds so this bird is not added again
                            self.set_all_birds()
                        fl_exotic = False
                        for item in self.exotic_guides_birds:
                            if item[0] == bird_id and item[1] == guide_id:
                                fl_exotic = True
                        if not fl_exotic:
                            target_data = self.get_targets(guide_id, scientific_name)
                            speciality_data = self.get_specialities(guide_id, scientific_name)
                            params = (bird_id, guide_id, res_status_id, target_data[1], target_data[0],
                                      speciality_data[0], speciality_data[1])
                            utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                     sp='sp_insert_exotic_checklist',
                                                     params=' @BirdID=?, @GuideID=?, @ResidentStatusID=?, @Target=?, '
                                                            '@Likelihood=?, @EndemicStatus=?, @ConservationStatus=?',
                                                     params_values=params)
                            utilities.run_sql_params()
                row_ct += 1


class EbirdBarchartParseUtility(GuideBase):
    def __init__(self, logger, ebird_base_url, abundance_matrix, sql_server_connection, file_path=None):
        self.ebird_base_url = ebird_base_url
        self.abundance_matrix = abundance_matrix
        self.sql_server_connection = sql_server_connection
        self.barchart_suffix = '&yr=all&m='
        self.table_target = "table class=\"barChart\""
        self.bird_row_target = "rC_Row"
        self.file_path = file_path
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, file_path=file_path)

    def parse_all_regions(self):
        self.logger.info('Start Execution.')
        self.set_clements()
        self.set_all_birds()
        self.set_regions_birds()
        utilities = SQLUtilities(sp='sp_get_ebird_region_codes', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values='', params='')
        regions = utilities.run_sql_return_no_params()
        bird_ids_added = []
        for region in regions:
            region_id = region[0]
            url = self.ebird_base_url + region[1] + self.barchart_suffix
            website = requests.get(url)
            soup = BeautifulSoup(website.content, 'html5lib')
            barchart_tables = soup.findAll("table", {"class": "barChart"})
            # Page has 2 tables with the class name skip first table
            my_table = barchart_tables[1]
            rows = my_table.findChildren(['tr'])
            row_ct = 0
            for row in rows:
                # skip first row of month labels
                td_ct = 0
                bird_name = ''
                flag_clement_match = False
                flag_all_birds_match = False
                code = None
                bird_id = None
                scientific = None
                skip_non_species = False
                scores = []
                if row_ct > 0:
                    week_num = 0
                    tds = row.find_all('td')
                    for td in tds:
                        td_ct += 1
                        # bird name is at first TD cell
                        if td_ct == 1:
                            # catch the non species birds as NoneType Attribute errors b/c there are no anchor tags
                            try:
                                bird_name = td.find('a').contents[0]
                            except AttributeError as err:
                                skip_non_species = True
                                continue
                        # 12 months are in rows 4-15
                        if td_ct >= 4 or td_ct <= 15:
                            if not skip_non_species and bird_name:
                                weeks = td.findAll("div")
                                for week in weeks:
                                    week_num += 1
                                    abundance = ''
                                    # get the 2 char class from the div tag (which has abundance number)
                                    raw_abun = str(week)[12:14]
                                    if raw_abun == 'sp':
                                        abundance = '0'
                                    if raw_abun[0] == 'b':
                                        # remove the leading b top get number 1-9
                                        abundance = raw_abun[1]
                                        if abundance == 'u':
                                            abundance = ''
                                    diction = {str(week_num): str(abundance)}
                                    scores.append(diction)
                if not skip_non_species and bird_name:
                    scores = json.dumps(scores)
                    bird_name = bird_name.strip()
                    # match this bird common name to Clements taxonomy
                    for taxon in self.get_clements():
                        if taxon[1] == bird_name.strip():
                            flag_clement_match = True
                            code = taxon[4]
                            scientific = taxon[2]
                    if not flag_clement_match:
                        # add this bird to error table and break out of loop and go to next bird
                        params = (bird_name, region_id, str(scores))
                        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                 sp='sp_insert_region_abundance_err',
                                                 params=' @BirdName=?,@RegionID=?,@WeeksScores=?',
                                                 params_values=params)
                        utilities.run_sql_params()
                        continue
                    # get bird id with common name
                    # if not in birds table add it and return id
                    for bird in self.all_birds:
                        if bird[1] == bird_name.strip():
                            flag_all_birds_match = True
                            bird_id = bird[0]
                    if not flag_all_birds_match:
                        params = (bird_name.strip(), code, scientific, 1006)
                        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                 sp='sp_insert_bird',
                                                 params=' @BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?',
                                                 params_values=params)
                        bird_id = utilities.run_sql_return_params()[0][0]
                        bird_ids_added.append(bird_id)
                        self.logger.info("Added a new bird to the Birds table: " + bird_name.strip() +
                                         ' ,id: ' + str(bird_id))
                        # refresh the all birds so this bird is not added again
                        self.set_all_birds()
                    # check to see if this combination of region and bird id is already in the database
                    # if already in update not not insert
                    fl_regions_birds = False
                    for item in self.regions_birds:
                        if item[0] == region_id and item[1] == bird_id:
                            fl_regions_birds = True
                    if not fl_regions_birds:
                        params = (bird_id, region_id, str(scores))
                        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                 sp='sp_insert_region_abundance',
                                                 params=' @BirdID=?,@RegionID=?,@WeeksScores=?',
                                                 params_values=params)
                        utilities.run_sql_params()
                    else:
                        params = (bird_id, region_id, str(scores))
                        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                                 sp='sp_update_region_abundance',
                                                 params=' @BirdID=?,@RegionID=?,@WeeksScores=?',
                                                 params_values=params)
                        utilities.run_sql_params()
                row_ct += 1
        # add the new bird added to the file.
        self.logger.info("End script execution.")

#class UpdateGuides(GuideBase):




class CreateGuide(GuideBase):
    def __init__(self, file_path, logger, sql_server_connection, guide_name, audio_path, image_path,
                 playlist_root, test, root, targets_only, exotic_name=None):
        self.exotic_name = exotic_name
        self.targets_only = targets_only
        self.guide_name = guide_name
        self.file_path = file_path
        self.audio_path = audio_path
        self.image_path = image_path
        self.playlist_root = playlist_root
        self.test = test
        self.root = root
        self.guide_name = guide_name
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, file_path=file_path)

    def _process_exotic_file(self):
        return_list = []
        file = 'Exotic_' + self.exotic_name.replace(' ', '_') + '.xlsx'
        wb = load_workbook(self.file_path + file)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[1]:
                flag = False
                code = None
                english = None
                for taxon in self.get_clements():
                    if taxon[2] == row[2]:
                        flag = True
                        code = taxon[4]
                        english = taxon[1]
                if flag:
                    diction = {'name': english, 'code': code, 'scientific': row[2], 'target': '', 'id': ''}
                    return_list.append(diction)
                else:
                    msg = "This exotic bird scientific name " + row[2] + " did not match Clements taxonomy, " \
                          "please fix before proceeding. English name: " + row[1]
                    self.logger.error(msg)
                    raise TaxonomyException(msg)
        return return_list

    def _get_targets(self):
        return_list = []
        file = 'Exotic_' + self.exotic_name.replace(' ', '_') + '_Targets.xlsx'
        wb = load_workbook(self.file_path + file)
        sheetname = "Sheet1"
        ws = wb[sheetname]
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row[1]:
                diction = {'name': row[1], 'scientific': row[2]}
                return_list.append(diction)
        return return_list

    def _combine_ebird_exotic(self, ebird_list, exotic_list):
        for exotic_bird in exotic_list:
            flag = False
            for ebird in ebird_list:
                if ebird['scientific'] == exotic_bird['scientific']:
                    flag = True
            if not flag:
                ebird_list.append(exotic_bird)
        return ebird_list

    def run_create(self):
        self.logger.info('Start script execution to create Bird Guide.')
        self.set_clements()
        self.set_guide_id()
        clements = self.get_clements()
        guide_id = self.get_guide_id()
        targets = None
        if self.exotic_name:
            targets = self._get_targets()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        all_birds = utilities.run_sql_return_no_params()

        # collect all the ebird region files into one list then remove duplicates
        all_ebird_data = self._process_ebird_files()

        # match ebird list to clements on english name and get scientific name and taxon code
        # if using the same clements version year there should be no logged mismatches
        ebird_list_clements = self._match_clements_ebird(clements, all_ebird_data)

        # get exotic birds list and match to clements on scientific name, get clements english name and taxon code
        # log birds that don't match and fix these manually
        exotic_birds_clements = None
        all_birds_clements = None
        if self.exotic_name:
            exotic_birds_clements = self._process_exotic_file()

        # if there are any exotic birds then add the diff to the ebird list (no duplicates)
        # but only do this if we have the boolean to only update Targets to NO
        if not self.targets_only:
            if exotic_birds_clements:
                all_birds_clements = self._combine_ebird_exotic(ebird_list_clements, exotic_birds_clements)
            else:
                all_birds_clements = ebird_list_clements
        else:
            all_birds_clements = ebird_list_clements

        # add targets
        if self.exotic_name:
            for target in targets:
                for bird in all_birds_clements:
                    if target['scientific'] == bird['scientific']:
                        bird['target'] = 'TARGET'

        # compare the full list to birds already in guides database
        # and create a final list for adding to the database
        add_list = []
        for final_bird in all_birds_clements:
            flag = False
            id = None
            for master_bird in all_birds:
                if master_bird[1] == final_bird['name']:
                    flag = True
                    id = master_bird[0]
            if not flag:
                add = 'ADD'
            else:
                add = ''
            diction = {'name': final_bird['name'], 'code': final_bird['code'], 'scientific': final_bird['scientific'],
                       'add': add, 'target': final_bird['target'], 'id': id}
            add_list.append(diction)
        new_image_list = []
        # create paths for new images and audio files
        guide_des = 'New_Guide ' + self.guide_name + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        image_path = self.image_path
        audio_path = self.audio_path + guide_des + "\\"
        if self._check_new_add(add_list):
            try:
                os.mkdir(self.audio_path + guide_des)
            except FileExistsError as err:
                pass
        for add in add_list:
            target = 0
            if add['target'] == "TARGET":
                target = 1
            if add['add'] == "ADD":
                # add to birds database
                params_values = (add['name'], add['code'], add['scientific'], 1006)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird',
                                         params='@BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?')
                if self.test:
                    bird_id = 0
                else:
                    bird_id = utilities.run_sql_return_params()
                if self.test:
                    add['id'] = bird_id
                else:
                    add['id'] = bird_id[0][0]
                self.logger.info("Added new bird to database: " + add['name'])
                # todo calculate the Residency and Endemic status by looking at adjacent regions guides for this bird
                params_values = (add['id'], guide_id, 1, 2, target, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                if self.test:
                    pass
                else:
                    utilities.run_sql_params()
                self.logger.info("Added bird " + add['name'] + ' to the guide: ' + self.guide_name)
                diction = {'name': add['code'] + ' ' + add['name'], 'scientific': add['scientific']}
                new_image_list.append(diction)
                shutil.copy(self.audio_path + 'blank.mp3', audio_path + add['code'] + ' ' + add['name'] + '.mp3')

        # add the birds already in the database to the guide
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_in_guide', params='@GuideID=?', params_values=guide_id)
        guide_birds = utilities.run_sql_return_params()
        for add in add_list:
            flag = False
            target = 0
            if add['target'] == "TARGET":
                target = 1
            for name in guide_birds:
                if add['name'] == name[1]:
                    flag = True
            if not flag:
                params_values = (add['id'], guide_id, 1, 2, target, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                if self.test:
                    pass
                else:
                    utilities.run_sql_params()
                self.logger.info("Added bird " + add['name'] + ' to the guide: ' + self.guide_name)

        if self._check_new_add(add_list):
            f = open(self.root + guide_des + '.csv', "w")
            for item in new_image_list:
                f.write('"' + item['name'] + '"' + ',"' + item['scientific'] + '"' + '\n')
            f.close()
            self.logger.info("Playlist updated.")

        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values=guide_id, params='@GuideID=?', sp='sp_update_guide_last_update')
        if self.test:
            pass
        else:
            utilities.run_sql_params()
        self.logger.info('End Script Execution.\n')


class UpdateGuide(GuideBase):
    def __init__(self, file_path, logger, sql_server_connection, guide_name, audio_path, image_path,
                 playlist_root, test, root):
        self.guide_name = guide_name
        self.audio_path = audio_path
        self.image_path = image_path
        self.playlist_root = playlist_root
        self.test = test
        self.root = root
        self.guide_name = guide_name
        GuideBase.__init__(self, logger=logger, sql_server_connection=sql_server_connection, file_path=file_path)

    def run_update(self):
        self.logger.info('Start script execution to update Bird Guide.')
        self.set_clements()
        self.set_guide_id()
        clements = self.get_clements()
        guide_id = self.get_guide_id()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_all_birds')
        all_birds = utilities.run_sql_return_no_params()
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 sp='sp_get_in_guide', params='@GuideID=?', params_values=guide_id)
        guide_birds = utilities.run_sql_return_params()

        # collect all the ebird region files into one list then remove duplicates
        all_ebird_data = self._process_ebird_files()

        # get clements data
        ebird_list_clements = self._match_clements_ebird(clements, all_ebird_data)

        # create paths for new images and audio files
        guide_des = 'Updates_Guide ' + self.guide_name + '_' + datetime.datetime.today().strftime('%Y-%m-%d')
        image_path = self.image_path
        audio_path = self.audio_path + guide_des + "\\"

        new_image_list = []
        # check for birds new to the birds database
        if self._check_new_update(ebird_list_clements, all_birds):
            try:
                os.mkdir(self.audio_path + guide_des)
            except FileExistsError as err:
                pass
        for ebird in ebird_list_clements:
            flag = False
            for name in all_birds:
                if ebird['name'] == name[1]:
                    ebird['id'] = name[0]
                    flag = True
            if not flag:
                print('Not in Birds Database:' + ebird['name'])
                # add to birds database
                params_values = (ebird['name'], ebird['code'], ebird['scientific'], 1006)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird',
                                         params='@BirdName=?,@TaxanomicCode=?,@ScientificName=?,@Artist=?')
                if self.test:
                    bird_id = 0
                else:
                    bird_id = utilities.run_sql_return_params()
                if self.test:
                    ebird['id'] = bird_id
                else:
                    ebird['id'] = bird_id[0][0]
                self.logger.info("Added new bird to database: " + ebird['name'])
                diction = {'name': ebird['code'] + ' ' + ebird['name'], 'scientific': ebird['scientific']}
                new_image_list.append(diction)
                shutil.copy(self.audio_path + 'blank.mp3', audio_path + ebird['code'] + ' ' + ebird['name'] + '.mp3')

        # check for birds new to the guide
        for ebird in ebird_list_clements:
            flag = False
            for name in guide_birds:
                if ebird['name'] == name[1]:
                    flag = True
            if not flag:
                print('Not in Birds Guide:' + ebird['name'])
                params_values = (ebird['id'], guide_id, 1, 2, 0, 5)
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params_values, sp='sp_insert_bird_guide',
                                         params='@BirdID=?,@GuideID=?,@ResidentID=?,@Difficulty=?,@Target=?,@Endemic=?')
                if self.test:
                    pass
                else:
                    utilities.run_sql_params()
                self.logger.info("Added bird " + ebird['name'] + ' to the guide: ' + self.guide_name)

        if self._check_new_update(ebird_list_clements, all_birds):
            f = open(self.root + guide_des + '.csv', "w")
            for item in new_image_list:
                f.write('"' + item['name'] + '"' + ',"' + item['scientific'] + '"' + '\n')
            f.close()
            self.logger.info("New birds added to file.")

        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values=guide_id, params='@GuideID=?', sp='sp_update_guide_last_update')
        if self.test:
            pass
        else:
            utilities.run_sql_params()
        self.logger.info("End script execution.")
