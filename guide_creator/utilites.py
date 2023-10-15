import pyodbc, os, datetime
from guide_creator.exceptions import DatabaseConnectionException, DatabaseOperationException
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from bs4 import BeautifulSoup
import requests, json, math
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class UtilitiesBase:
    def __init__(self, logger):
        self.logger = logger

    def parse_abundance(self, str_abundance):
        return_string = []
        for myChar in str_abundance:
            if myChar.strip():
                return_string.append(myChar)
        return return_string

    def connect_sqlserver(self, sqlserver_connection):
        connection_string = "Driver={ODBC Driver 17 for SQL Server};"
        connection_string += "Server=" + sqlserver_connection.name + ";"
        connection_string += "Database=" + sqlserver_connection.database + ";"
        connection_string += "Trusted_Connection=yes;"
        try:
            conn = pyodbc.connect(connection_string)
        except (pyodbc.OperationalError, pyodbc.InterfaceError) as err:
            msg = 'Error connecting to database via odbc: ' + str(err)
            self.logger.error(msg)
            raise DatabaseConnectionException(msg)
        return conn


class SQLUtilities(UtilitiesBase):
    def __init__(self, logger, sql_server_connection, sp=None, sql=None, params=None, params_values=None):
        self.params = params
        self.params_values = params_values
        self.sql_server_connection = sql_server_connection
        self.sp = sp
        self.sql = sql
        UtilitiesBase.__init__(self, logger=logger)

    def run_sql_return_params_no_sp(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        my_params = self.params_values
        try:
            cursor.execute(self.sql, my_params)
            data = cursor.fetchall()
        except pyodbc.ProgrammingError as err:
            msg = "An error occurred executing a sql server command"
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()
        return data

    def run_sql_return_no_params(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        sql = "EXEC " + self.sp + ';'
        try:
            cursor.execute(sql)
            data = cursor.fetchall()
        except pyodbc.ProgrammingError as err:
            msg = "An error occurred executing a sql server command"
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()
        return data

    def run_sql_return_params(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        sql = "EXEC " + self.sp + ' ' + self.params + ';'
        my_params = self.params_values
        try:
            cursor.execute(sql, my_params)
            data = cursor.fetchall()
        except pyodbc.ProgrammingError as err:
            msg = "An error occurred executing a sql server command"
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()
        return data

    def run_sql_params(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        sql = "EXEC " + self.sp + ' ' + self.params + ';'
        my_params = self.params_values
        try:
            cursor.execute(sql, my_params)
        except (pyodbc.IntegrityError, pyodbc.ProgrammingError) as err:
            msg = "An error occurred executing a sql server command, message: " + str(err)
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()

    def run_sql(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        sql = "EXEC " + self.sp + ';'
        try:
            cursor.execute(sql)
        except pyodbc.ProgrammingError as err:
            msg = "An error occurred executing a sql server command"
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()

    def run_plain_sql_return(self):
        conn = self.connect_sqlserver(self.sql_server_connection)
        cursor = conn.cursor()
        sql = self.sql
        try:
            cursor.execute(sql)
            data = cursor.fetchall()
        except pyodbc.ProgrammingError as err:
            msg = "An error occurred executing a sql server command"
            raise DatabaseOperationException(msg)
        conn.commit()
        cursor.close()
        conn.close()
        return data


class GoogleAPIUtilities(UtilitiesBase):
    def __init__(self, logger, scopes, cred_path, drive_root):
        self.scopes = scopes
        self.cred_path = cred_path
        self.drive_root = drive_root
        UtilitiesBase.__init__(self, logger=logger)

    def authenticate(self):
        credentials = service_account.Credentials.from_service_account_file(self.cred_path + 'credentials.json',
                                                                            scopes=self.scopes)
        return build('drive', 'v3', credentials=credentials)

    def list_folders_id_by_name(self, service):
        files = service.files().list(q="mimeType='application/vnd.google-apps.folder' and name='" + self.drive_root + "'",
                                     spaces='drive', fields='nextPageToken, files(id, name)').execute()
        # there should only be one but if more this gets the first on
        return files['files'][0]['id']

    def list_all_folders_py_parent(self, service, file_id):
        folders = service.files().list(q="mimeType='application/vnd.google-apps.folder' and '" +
                                       file_id + "' in parents", spaces='drive',
                                       fields='nextPageToken, files(id, name)').execute()
        return folders

    def list_permissions_by_file_id(self, service, file_id):
        permissions = service.permissions().list(fileId=file_id, fields='*').execute()
        return permissions

    def delete_file_or_directory(self, service, file_id):
        service.files().delete(fileId=file_id).execute()

    def create_file_or_directory(self, service, item_name, parent_id):
        file_metadata = {
            'name': item_name,
            'parents': [parent_id],
            'mimeType': 'application/vnd.google-apps.folder'
        }
        file = service.files().create(body=file_metadata, fields='id').execute()
        return file['id']

    def create_document_upload(self, service, doc_name, doc_path, parent_id, mimetype, meta_mine_type):
        file_metadata = {
            'name': doc_name,
            'mimeType': meta_mine_type,
            'parents': [parent_id]
        }
        media = MediaFileUpload(doc_path + doc_name, mimetype="'" + mimetype + "'", resumable=True)
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return file['id']

    def create_media_upload(self, service, media_name, media_path, parent_id, mimetype):
        file_metadata = {
            'name': media_name,
            'parents': [parent_id]
        }
        media = MediaFileUpload(media_path + media_name,
                                mimetype="'" + mimetype + "'",
                                resumable=True)
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        return file['id']

    def create_permission(self, service, file_id, email):
        file_metadata = {
            'role': 'reader',
            'type': 'user',
            'emailAddress': email
        }
        permission = service.permissions().create(body=file_metadata, fileId=file_id).execute()
        return permission


class GoogleDriveSuperGuide:
    def __init__(self, logger, sql_server_connection, audio_path, google_api_scopes, google_cred_path,
                 root_guide_dir, super_guide_id, super_guide_name, super_guide_perm):
        self.logger = logger
        self.sql_server_connection = sql_server_connection
        self.audio_path = audio_path
        self.google_api_scopes = google_api_scopes
        self.google_cred_path = google_cred_path
        self.root_guide_dir = root_guide_dir
        self.super_guide_id = super_guide_id
        self.super_guide_name = super_guide_name
        self.super_guide_perm = super_guide_perm

    def refresh(self):
        google_api = GoogleAPIUtilities(self.logger, self.google_api_scopes, self.google_cred_path,
                                        self.root_guide_dir)
        service = google_api.authenticate()
        # get root directory id where all the bird directories will be found
        root_id = google_api.list_folders_id_by_name(service=service)
        # todo check for existence of any folder here if exists abort and info message to remove all directories
        # create the directory
        new_folder_id = google_api.create_file_or_directory(service=service, item_name=self.super_guide_name,
                                                            parent_id=root_id)
        emails = []
        # todo check no match superguide name if not match abort and info to add superguide to config
        for item in self.super_guide_perm:
            if item['Superguide'] == self.super_guide_name:
                for email in item['Emails']:
                    emails.append(email)
        if emails:
            for email in emails:
                new_perm_id = google_api.create_permission(service=service, file_id=new_folder_id, email=email)
        utilities = SQLUtilities(sp='sp_get_birds_in_super_guide', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values=self.super_guide_id, params='@SuperGuideID=?')
        birds = utilities.run_sql_return_params()
        # upload audio files
        for bird in birds:
            google_api.create_media_upload(service=service, media_name=bird[0] + '.mp3', media_path=self.audio_path,
                                           parent_id=new_folder_id, mimetype='image/jpeg')


class AbundanceChartSuperGuide(GoogleAPIUtilities):
    def __init__(self, logger, sql_server_connection, drive_root, super_guide_id, super_guide_name,
                 super_guide_perm, chart_root, google_api_scopes=None, google_cred_path=None):
        self.sql_server_connection = sql_server_connection
        self.chart_root = chart_root
        self.super_guide_id = super_guide_id
        self.super_guide_name = super_guide_name
        self.super_guide_perm = super_guide_perm
        GoogleAPIUtilities.__init__(self, logger=logger, drive_root=drive_root, cred_path=google_cred_path,
                                    scopes=google_api_scopes)

    def refresh(self):
        super_guide_chart_path = self.chart_root
        utilities = SQLUtilities(sp='sp_get_abundance_updated_date', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection)
        abundance_date = utilities.run_sql_return_no_params()[0][0]
        today_date = datetime.datetime.today().strftime('%Y-%m-%d')
        if not os.path.exists(super_guide_chart_path):
            os.mkdir(super_guide_chart_path)
        emails = []
        google_api = GoogleAPIUtilities(self.logger, self.scopes, self.cred_path,
                                        drive_root=self.drive_root)
        service = google_api.authenticate()
        root_id = google_api.list_folders_id_by_name(service=service)
        folders = google_api.list_all_folders_py_parent(service=service, file_id=root_id)
        permissions = None
        file_id = None
        for folder in folders['files']:
            if self.super_guide_name == folder['name']:
                file_id = folder['id']
                permissions = google_api.list_permissions_by_file_id(service=service, file_id=folder['id'])
        # then delete directory and all files
        if file_id:
            google_api.delete_file_or_directory(service=service, file_id=file_id)
        # create the directory
        new_folder_id = google_api.create_file_or_directory(service=service, item_name=self.super_guide_name,
                                                            parent_id=root_id)
        if permissions:
            for perm in permissions['permissions']:
                if perm['role'] != 'owner':
                    emails.append(perm['emailAddress'])
        if emails:
            # now get a list of active guides in this superguide and create directory for each
            utilities = SQLUtilities(sp='sp_get_active_guides_in_super_guide', logger=self.logger,
                                     params_values=self.super_guide_id, params='@SuperGuideID=?',
                                     sql_server_connection=self.sql_server_connection)
            guides = utilities.run_sql_return_params()
            for guide in guides:
                chart_path = super_guide_chart_path
                utilities = SQLUtilities(sp="sp_get_pl_guide", logger=self.logger,
                                         sql_server_connection=self.sql_server_connection, params_values=guide[1],
                                         params='@GuideID=?')
                return_values = utilities.run_sql_return_params()
                # make the first chart with abundance in one column and residence, endemic, and conservation
                chart_name = guide[2] + ' Abundance Chart'
                book = Workbook()
                sheet = book.active
                birds = []
                for item in return_values:
                    my_data = [item[1], item[4], item[5], item[6], item[7]]
                    birds.append(my_data)
                # add empty line in chart
                birds.append(['', '', '', '', ''])
                # add footer in chart
                birds.append(['Abundance Refresh: ' + str(abundance_date), 'Chart Refresh: '
                              + str(today_date), '', '', '', ''])
                header = ['Species', 'Ebird Abundance', 'Residency', 'Endemic PH', 'Conservation']
                sheet.append(header)
                sheet.append([])
                for bird in birds:
                    sheet.append(bird)
                for column_cells in sheet.columns:
                    new_column_letter = (get_column_letter(column_cells[0].column))
                    if new_column_letter == 'A':
                        sheet.column_dimensions[new_column_letter].width = 30
                    elif new_column_letter == 'B':
                        sheet.column_dimensions[new_column_letter].width = 18
                    elif new_column_letter == 'C':
                        sheet.column_dimensions[new_column_letter].width = 23
                    elif new_column_letter == 'D':
                        sheet.column_dimensions[new_column_letter].width = 12
                    elif new_column_letter == 'E':
                        sheet.column_dimensions[new_column_letter].width = 20
                book.save(chart_path + '\\' + chart_name + '.xlsx')
                google_api.create_document_upload(service=service, doc_name=chart_name + '.xlsx',
                                                  doc_path=chart_path, parent_id=new_folder_id,
                                                  mimetype='text/csv',
                                                  meta_mine_type='application/vnd.google-apps.document')
                # make the second chart with abundance per month and residence
                chart_name = guide[2] + ' Abundance Detail Chart'
                book = Workbook()
                sheet = book.active
                birds = []
                for item in return_values:
                    abundance_list = self.parse_abundance(item[4])
                    if not abundance_list:
                        abundance_list = ['', '', '', '', '', '', '', '', '', '', '', '']
                    my_data = [item[1], abundance_list[0], abundance_list[1], abundance_list[2],
                               abundance_list[3], abundance_list[4], abundance_list[5], abundance_list[6],
                               abundance_list[7], abundance_list[8], abundance_list[9], abundance_list[10],
                               abundance_list[11], item[5]]
                    birds.append(my_data)
                # add empty line in chart
                birds.append(['', '', '', '', ''])
                # add footer in chart
                birds.append(['Abundance Refresh: ' + str(abundance_date), 'Chart Refresh: '
                              + str(today_date), '', '', '', ''])
                header = ['Species', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct',
                          'Nov', 'Dec', 'Residency']
                sheet.append(header)
                sheet.append([])
                for bird in birds:
                    sheet.append(bird)
                for column_cells in sheet.columns:
                    new_column_letter = (get_column_letter(column_cells[0].column))
                    if new_column_letter == 'A':
                        sheet.column_dimensions[new_column_letter].width = 30
                    elif new_column_letter == 'B':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'C':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'D':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'E':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'F':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'G':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'H':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'I':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'J':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'K':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'L':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'M':
                        sheet.column_dimensions[new_column_letter].width = 4
                    elif new_column_letter == 'N':
                        sheet.column_dimensions[new_column_letter].width = 20
                book.save(chart_path + '\\' + chart_name + '.xlsx')
                google_api.create_document_upload(service=service, doc_name=chart_name + '.xlsx',
                                                  doc_path=chart_path, parent_id=new_folder_id,
                                                  mimetype='text/csv',
                                                  meta_mine_type='application/vnd.google-apps.document')


class PlaylistsSuperGuide(GoogleAPIUtilities):
    def __init__(self, logger, sql_server_connection, playlist_root, drive_root, super_guide_id, super_guide_name,
                 super_guide_perm, google_api_scopes=None, google_cred_path=None):
        self.sql_server_connection = sql_server_connection
        self.playlist_root = playlist_root
        self.super_guide_id = super_guide_id
        self.super_guide_name = super_guide_name
        self.super_guide_perm = super_guide_perm
        GoogleAPIUtilities.__init__(self, logger=logger, drive_root=drive_root, cred_path=google_cred_path,
                                    scopes=google_api_scopes)

    def month_list_include(self, months, abundance):
        include = False
        c = 0
        for item in abundance:
            c += 1
            for m in months:
                if c == m[0]:
                    # anything but no observations (-) and no data (*) is considered True
                    if item.strip() != '-' and item.strip() != '*':
                        include = True
        return include

    def refresh(self):
        super_guide_playlist_path = self.playlist_root + self.super_guide_name + '\\'
        if not os.path.exists(super_guide_playlist_path):
            os.mkdir(super_guide_playlist_path)
        # now get a list of active guides in this superguide and create directory for each
        utilities = SQLUtilities(sp='sp_get_active_guides_in_super_guide', logger=self.logger,
                                 params_values=self.super_guide_id, params='@SuperGuideID=?',
                                 sql_server_connection=self.sql_server_connection)
        guides = utilities.run_sql_return_params()
        for guide in guides:
            playlists_sps = [{'sp': 'sp_get_pl_breeding', 'name': 'Breeding'},
                             {'sp': 'sp_get_pl_common', 'name': 'Common'},
                             {'sp': 'sp_get_pl_common_passerines', 'name': 'Passerines C-A'},
                             {'sp': 'sp_get_pl_common_scarce_passerines', 'name': 'Passerines s-A'},
                             {'sp': 'sp_get_pl_common_uncommon_doves', 'name': 'Doves Cuckoos UC-A'},
                             {'sp': 'sp_get_pl_doves', 'name': 'Doves Cuckoos'},
                             {'sp': 'sp_get_pl_guide', 'name': ''},
                             {'sp': 'sp_get_pl_hawks', 'name': 'Hawks Owls'},
                             {'sp': 'sp_get_pl_herons', 'name': 'Rails Herons'},
                             {'sp': 'sp_get_pl_kingfishers', 'name': 'Kingfihsers'},
                             {'sp': 'sp_get_pl_passerines', 'name': 'Passerines'},
                             {'sp': 'sp_get_pl_seabirds', 'name': 'Seabirds'},
                             {'sp': 'sp_get_pl_shorebirds', 'name': 'Shorebirds'},
                             {'sp': 'sp_get_pl_swifts', 'name': 'Swifts Nightjars'},
                             {'sp': 'sp_get_pl_winter', 'name': 'Winter'}]
            header = '#EXTM3U\n'
            item_begin = '#EXTINF:'
            extension = '.mp3\n'
            # todo config this or make variable based on super guide?
            folder_phone = 'Philippines/'
            # todo get phone root from database
            root_phone = '/storage/emulated/0/'
            playlist_path = super_guide_playlist_path + guide[0]
            if not os.path.exists(playlist_path):
                os.mkdir(playlist_path)
            else:
                for fil in os.listdir(playlist_path):
                    os.remove(os.path.join(playlist_path, fil))
            for item in playlists_sps:
                if item['name'] == '':
                    playlist_name = guide[2]
                else:
                    playlist_name = guide[2] + ' ' + item['name']
                str_sp = item['sp']
                utilities = SQLUtilities(sp=str_sp, logger=self.logger,
                                         sql_server_connection=self.sql_server_connection,
                                         params_values=guide[1], params='@GuideID=?')
                birds = utilities.run_sql_return_params()
                str_file = header
                for bird in birds:
                    str_file += item_begin
                    str_file += str(bird[2]) + ',' + bird[3] + " - "
                    str_file += bird[0] + ' ' + bird[1] + '\n'
                    str_file += root_phone + folder_phone + bird[0] + ' ' + bird[1] + extension
                file_path = playlist_path + "\\" + playlist_name + '.m3u'
                with open(file_path, 'w') as myfile:
                    myfile.write(str_file)
                    myfile.close()
                # now create the months based playlist version if months are configured in DB
                utilities = SQLUtilities(sp='sp_get_months_playlist_data', logger=self.logger,
                                         sql_server_connection=self.sql_server_connection)
                months = utilities.run_sql_return_no_params()
                if months:
                    months_suffix = ''
                    for m in months:
                        months_suffix += m[1].strip() + '_'
                    months_suffix = months_suffix[:-1]
                    file_path = playlist_path + "\\" + playlist_name + '_' + months_suffix + '.m3u'
                    str_file = header
                    for bird in birds:
                        # check to see if the abundance string has a letter (r,s,C,U,A) in the months positions,
                        # if not don't include in the playlist
                        abundance_raw = bird[4]
                        if abundance_raw:
                            abundance = self.parse_abundance(bird[4])
                            # needs to be truthy and not a hyphen in at least one of the months
                            if self.month_list_include(months, abundance):
                                str_file += item_begin
                                str_file += str(bird[2]) + ',' + bird[3] + " - "
                                str_file += bird[0] + ' ' + bird[1] + '\n'
                                str_file += root_phone + folder_phone + bird[0] + ' ' + bird[1] + extension
                        with open(file_path, 'w') as myfile:
                            myfile.write(str_file)
                            myfile.close()


class ParseEbirdRegions(UtilitiesBase):
    def __init__(self, logger, country, ebird_base_url, sql_server_connection, counties):
        self.country = country
        self.counties = counties
        self.ebird_base_url = ebird_base_url
        self.suffix = '/regions?yr=all&m='
        self.sql_server_connection = sql_server_connection
        UtilitiesBase.__init__(self, logger=logger)

    def _enter_region(self, item_name, item_code):
        # if country and region code don't exist in regions table row add it now
        utilities = SQLUtilities(sp='sp_get_region', logger=self.logger,
                                 sql_server_connection=self.sql_server_connection,
                                 params_values=(item_code, self.country),
                                 params=' @RegionCode=?, @Country=?')
        region = utilities.run_sql_return_params()
        if not region:
            utilities = SQLUtilities(sp='sp_insert_ebird_region', logger=self.logger,
                                     sql_server_connection=self.sql_server_connection,
                                     params_values=(self.country, item_code, item_name),
                                     params=' @Country=?, @RegionCode=?, @RegionName=?')
            utilities.run_sql_params()
            self.logger.info("Added a new Ebird Region for " + self.country + ": " + item_name)

    def run(self):
        self.logger.info('Start script execution.')
        url = self.ebird_base_url + self.country + self.suffix
        website = requests.get(url)
        soup = BeautifulSoup(website.content, 'html5lib')
        table = soup.find("table", {"class": "Table Table--noBorder Table--clearRows"})
        rows = table.findChildren('tr')
        c = 0
        row_ct = 0
        for row in rows:
            if row_ct > 0:
                td_ct = 0
                tds = row.find_all('td')
                for td in tds:
                    if td_ct == 1:
                        link = td.find('a')
                        region_name = link.contents[0]
                        href = link.attrs['href']
                        href_spl = href.split('?')
                        region_code = href_spl[0][14:]
                        if self.counties:
                            url = self.ebird_base_url + region_code + self.suffix
                            website = requests.get(url)
                            soup = BeautifulSoup(website.content, 'html5lib')
                            table = soup.find("table", {"class": "Table Table--noBorder Table--clearRows"})
                            rows = table.findChildren('tr')
                            ct_row_ct = 0
                            for row in rows:
                                if ct_row_ct > 0:
                                    ct_td_ct = 0
                                    tds = row.find_all('td')
                                    for td in tds:
                                        if ct_td_ct == 1:
                                            c += 1
                                            link = td.find('a')
                                            county_name = link.contents[0]
                                            href = link.attrs['href']
                                            href_spl = href.split('?')
                                            county_code = href_spl[0][14:]
                                            self._enter_region(county_name, county_code)
                                        ct_td_ct += 1
                                ct_row_ct += 1
                        else:
                            self._enter_region(region_name, region_code)
                    td_ct += 1
            row_ct += 1
        self.logger.info('End script execution.')


class ParseGuideAbundance(UtilitiesBase):
    def __init__(self, logger, sql_server_connection, ebird_matrix):
        self.sql_server_connection = sql_server_connection
        self.ebird_matrix = ebird_matrix
        UtilitiesBase.__init__(self, logger=logger)

    def _translate_regions_abundance(self, raw_data):
        return_data = ''
        for myData in raw_data:
            for item in self.ebird_matrix:
                for key, val in item.items():
                    # insufficient data stored as empty string
                    if myData != '':
                        if val[0] < int(myData) <= val[1]:
                            return_data = return_data + key
            # has sufficient data but no observations stored as 0, show hyphen and space
            if myData == '0':
                return_data = return_data + '- '
            # insufficient data, stored as empty string, show asterisk
            if myData == '':
                return_data = return_data + '*'
        return return_data

    def _get_difficulty_id(self, abundance_string):
        # this takes the abundance data from ebird and determines an average/overall difficulty ID
        # need two months or more with abundant to be abundant (1)
        # need two or more months with common and 1 or less with abundant to be abundant, etc...
        ct_a = abundance_string.count('A')
        ct_c = abundance_string.count('C')
        ct_u = abundance_string.count('U')
        ct_s = abundance_string.count('s')
        ct_r = abundance_string.count('r')
        if ct_a >= 2:
            return_int = 1
        elif (ct_c >= 2 or ct_a == 1) and ct_a <= 1:
            return_int = 2
        elif (ct_u >= 2 or ct_c == 1) and ct_c <= 1 and ct_a == 0:
            return_int = 3
        elif (ct_s >= 2 or ct_u == 1) and ct_u <= 1 and ct_c == 0 and ct_a == 0:
            return_int = 4
        elif (ct_r >= 1 or ct_s == 1) and ct_s <= 1 and ct_u == 0 and ct_c == 0 and ct_a == 0:
            return_int = 5
        else:
            return 0
        return return_int

    def _process_avg_abundance(self, abundance_list):
        counts_by_month = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        sums_by_month = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        avg_abundance = []
        for region in abundance_list:
            # m is the index for the running counts and sums lists
            m = 0
            for month in region:
                # empty string means no abundance data available so don't include this in the running counts by month
                # nor in the running sum by month
                if month != '':
                    counts_by_month[m] += 1
                    sums_by_month[m] += int(month)
                m += 1
        for x in range(12):
            if counts_by_month[x] == 0:
                avg_abundance.append('0')
            else:
                # for guides with more than 4 regions, add a correction to the sum for all > 0 values
                if len(abundance_list) > 4:
                    average = sums_by_month[x] / counts_by_month[x]
                    if average > 0:
                        average += 0.6
                    avg_abundance.append(str(math.ceil(average)))
                else:
                    avg_abundance.append(str(math.ceil(sums_by_month[x] / counts_by_month[x])))
        return avg_abundance

    def calculate_region_abundance(self, bird_id, guide_id):
        # return series of 12 strings to represent the abundance of this bird in the guide for each month
        # for example:   -*-rsUCCs-- one for each month: JFMAMJJASOND
        # hpyhen = no observations (add a space after for readability),
        # * = insufficient data
        # max of each week in each month
        # AVG of all regions if > 1  region in the guide
        lst_regions_abundance = []
        params = (bird_id, guide_id)
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values=params, sp='sp_get_abundance_data',
                                 params='@BirdID=?, @GuideID=?')
        regions_abundance = utilities.run_sql_return_params()
        abundance_raw = []
        if regions_abundance:
            for region in regions_abundance:
                abundance_raw_list = []
                # convert stored json string to json object
                big_list = json.loads(region[2])
                # split into 12 moth groups of 4 weeks each
                n = 4
                lists_of_lists = [big_list[i:i + n] for i in range(0, len(big_list), n)]
                for four_list in lists_of_lists:
                    # create a list of the four values in each month and find the maximum value
                    this_item = []
                    for dict in four_list:
                        for key, value in dict.items():
                            this_item.append(value)
                    my_result = max(this_item)
                    abundance_raw_list.append(my_result)
                trans_data = self._translate_regions_abundance(abundance_raw_list)
                diction = {'region': region[1], 'country': region[0], 'data': trans_data, 'region_id': region[3]}
                lst_regions_abundance.append(diction)
                abundance_this_region = []
                for four_list in lists_of_lists:
                    # create a list of the four values in each month and find the maximum value
                    this_item = []
                    for dict in four_list:
                        for key, value in dict.items():
                            this_item.append(value)
                    # my_result is the max of each month in each region
                    my_result = max(this_item)
                    # list of the abundance for each month for this region
                    abundance_this_region.append(my_result)
                # this list is for all regions
                abundance_raw.append(abundance_this_region)
            if len(abundance_raw) == 1:
                abundance_final = abundance_raw[0]
            else:
                # calc the average of all the regions lists into one list
                abundance_final = self._process_avg_abundance(abundance_raw)
            str_regions_abundance = self._translate_regions_abundance(abundance_final)
            return [str_regions_abundance, lst_regions_abundance, abundance_final]
        else:
            return['', '', '']

    def update_abundance_calc(self):
        self.logger.info("Start script execution.")
        utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                 params_values='', sp='sp_get_all_birds_guides_ids',
                                 params='')
        ids = utilities.run_sql_return_no_params()
        for id in ids:
            abundance_data = self.calculate_region_abundance(bird_id=id[0], guide_id=id[1])
            if abundance_data[0]:
                params = (id[1], id[0], abundance_data[0])
                utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                         params_values=params, sp='sp_update_guide_abundance_string',
                                         params=' @GuideID=?, @BirdID=?, @AbundanceString=?')
                utilities.run_sql_params()
                for region in abundance_data[1]:
                    params = (region['region_id'], id[0], region['data'])
                    utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                             params_values=params, sp='sp_update_regions_abundance_string',
                                             params=' @RegionID=?, @BirdID=?, @AbundanceString=?')
                    utilities.run_sql_params()
                # calculate the difficulty ID from abundance data
                difficulty_id = self._get_difficulty_id(abundance_data[0])
            else:
                # no ebird data, make the difficulty ID = 7 (no ebird data)
                difficulty_id = 7
            params = (id[1], id[0], difficulty_id)
            # in stored procedure if difficulty is 6 it stays 6 otherwise it updates it
            utilities = SQLUtilities(logger=self.logger, sql_server_connection=self.sql_server_connection,
                                     params_values=params, sp='sp_update_bird_guide_difficulty',
                                     params=' @GuideID=?, @BirdID=?, @DifficultyID=?')
            utilities.run_sql_params()
