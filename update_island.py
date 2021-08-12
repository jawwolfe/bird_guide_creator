from openpyxl import load_workbook
import pyodbc, csv

connection_string = "Driver={ODBC Driver 17 for SQL Server};"
connection_string += "Server=localhost\\SQLEXPRESS;"
connection_string += "Database=BirdGuide;"
connection_string += "Trusted_Connection=yes;"
conn = pyodbc.connect(connection_string)
path_clements = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Clements_2019.xlsx'
path1 = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Cebu Ebird.xlsx'
path2 = ''


def get_bird_name(name):
    sql = "Select TaxanomicCode from Birds where BirdName = ?;"
    cursor = conn.cursor()
    params = name
    cursor.execute(sql,params)
    taxcode = cursor.fetchall()
    conn.commit()
    if taxcode:
        return (True, taxcode[0][0])
    else:
        return (False, '')


def get_bird_in_island(name, islandid):
    sql = "SELECT [ID] FROM [BirdsIslands] ISL inner join Birds BR on BR.BirdID = ISL.BirdID where BR.BirdName = ? and ISL.IslandID = ?;"
    cursor = conn.cursor()
    params = (name, islandid)
    cursor.execute(sql,params)
    mydata = cursor.fetchall()
    conn.commit()
    if mydata:
        return True
    else:
        return False


def process_ebird_file(path):
    my_wb = load_workbook(path)
    my_sheetname = "Sheet1"
    my_ws = my_wb[my_sheetname]
    raw_list = []
    for my_row in my_ws.iter_rows(min_row=2, values_only=True):
        raw_list.append(my_row[0])
    return_list = []
    for my_item in raw_list:
        # todo add another line to ignore the new page line characters
        if 'sp.' not in my_item:
            if '/' not in my_item:
                if 'Domestic' not in my_item:
                    if my_item != 'Jan':
                        return_list.append(my_item.replace('\t', ''))
    return return_list


wb = load_workbook(path_clements)
sheetname = "eBird Clements v2019 Aug 2019"
ws = wb[sheetname]
clements_data = []
clements_species = []
for row in ws.iter_rows(min_row=2, values_only=True):
     data = {'category': row[2], 'english': row[3], 'scientific': row[4]}
     clements_data.append(data)
for item in clements_data:
    if item['category'] == "species":
        bird = {'english': item['english'], 'scientific': item['scientific']}
        clements_species.append(bird)

lst_1 = process_ebird_file(path1)
if path2:
    lst_2 = process_ebird_file(path2)
else:
    lst_2 = []

all_ebird_birds = list(set(lst_1 + lst_2))

add_list = []
for bird in all_ebird_birds:
    code = ''
    scien = ''
    add = 'NO'
    flag_scien = False
    flag_database = False
    flag_island = False
    for taxon in clements_species:
        if taxon['english'].strip() == bird.strip():
            flag_scien = True
            scien = taxon['scientific']
    result = get_bird_name(bird)
    code = result[1]
    flag_database = result[0]
    if not flag_database:
        add = 'ADD'
    if flag_database:
        flag_island = get_bird_in_island(bird, 2)
    if not flag_island or not flag_database:
        if not flag_scien:
            print('Bird not found in master checklist!!!')
        else:
            diction = {'code': code, 'name': bird, 'scientific': scien, 'add': add, 'target': 'NO'}
            add_list.append(diction)
            print(diction)

f = open("cebu_add.csv", "w")
writer = csv.DictWriter(f, fieldnames=["code", "name", "scientific", "add", "target"], lineterminator='\n')
writer.writerows(add_list)
f.close()