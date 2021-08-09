from openpyxl import load_workbook
import csv
path3 = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\panay_all.xlsx'
path4 = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\panay_all_old.xlsx'


def process_ebird_file(path):
    wb = load_workbook(path)
    sheetname = "Sheet1"
    ws = wb[sheetname]
    return_list = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        data = {'name': row[1], 'code': row[0]}
        return_list.append(data)
    return return_list


lst_3 = process_ebird_file(path3)
lst_4 = process_ebird_file(path4)


add = []
for item in lst_3:
    flag = False
    for item2 in lst_4:
        if item['name'] == item2['name']:
            flag = True
    if not flag:
        diction = {'code': item['code'], 'name': item['name']}
        add.append(diction)


f = open("panay_add.csv", "w")
writer = csv.DictWriter(f, fieldnames=["code", "name"], lineterminator='\n')
writer.writerows(add)
f.close()
