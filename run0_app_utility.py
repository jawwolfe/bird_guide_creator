from guide_creator.utilities_guide import Compare, Rename, RecreateImageList, RecreateAudioFiles
from guide_creator.utilites import PlaylistsSuperGuide, SQLUtilities
from guide_creator.configs import config
from globals import initialize_sqlserver
from globals import initialize_logger, initialize_sqlserver
from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

SQLSERVER_NAME = config.SQLSERVER_NAME
SQLSERVER_DATABASE = config.SQLSERVER_DATABASE
AUDIO_PATH = config.AUDIO_PATH_GUIDE
IMAGE_PATH = config.IMAGE_PATH_GUIDE
PLAYLIST_ROOT = config.PLAYLIST_PATH
LOGGER = initialize_logger('bird_guide')
GOOGLE_API_SCOPES = config.GOOGLE_API_SCOPES
GOOGLE_CRED_PATH = config.GOOGLE_CRED_PATH


chart_path = 'C:\\temp'
chart_name =  'Cebu' + ' Abundance Chart'
book = Workbook()
sheet = book.active
birds = []
utilities = SQLUtilities(sp="sp_get_pl_guide", logger=LOGGER,
                         sql_server_connection=initialize_sqlserver(), params_values=2,
                         params='@GuideID=?')
return_values = utilities.run_sql_return_params()

for item in return_values:
    my_data = [item[1], item[4], item[5], item[6], item[7]]
    birds.append(my_data)
header = ['Species', 'Ebird Abundance', 'Residency', 'Endemic PH', 'Conservation']
sheet.append(header)
sheet.append([])
for bird in birds:
    sheet.append(bird)

for column_cells in sheet.columns:
    new_column_length = max(len(str(cell.value)) for cell in column_cells)
    new_column_letter = (get_column_letter(column_cells[0].column))
    if new_column_letter == 'A':
        sheet.column_dimensions[new_column_letter].width = 30
    elif new_column_letter == 'B':
        sheet.column_dimensions[new_column_letter].width = 18
    elif new_column_letter == 'C':
        sheet.column_dimensions[new_column_letter].width = 23
    elif new_column_letter == 'D':
        sheet.column_dimensions[new_column_letter].width = 12
    elif new_column_letter == 'E':
        sheet.column_dimensions[new_column_letter].width = 20

book.save(chart_path + '\\' + chart_name + '.xlsx')
