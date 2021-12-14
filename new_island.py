import pyodbc, csv, sys
from openpyxl import load_workbook


connection_string = "Driver={ODBC Driver 17 for SQL Server};"
connection_string += "Server=localhost\\SQLEXPRESS;"
connection_string += "Database=BirdGuide;"
connection_string += "Trusted_Connection=yes;"
conn = pyodbc.connect(connection_string)

path_exotic = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Palawan Exotic.xlsx'
path_clements = '/data/Clements_2019.xlsx'
path_targets = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Palawan Targets.xlsx'
path1 = 'C:\\Users\\Andrew\\PycharmProjects\\audioembedder\\Palawan Ebird.xlsx'
path2 = ''
path3 = ''

def process_ebird_file(path):
    my_wb = load_workbook(path)
    my_sheetname = "Sheet1"
    my_ws = my_wb[my_sheetname]
    raw_list = []
    for my_row in my_ws.iter_rows(min_row=2, values_only=True):
        raw_list.append(my_row[0])
    return_list = []
    for my_item in raw_list:
        # todo add another line to ignore the new page line characters
        if 'sp.' not in my_item:
            if '/' not in my_item:
                if 'Domestic' not in my_item:
                    if my_item != 'Jan':
                        return_list.append(my_item.replace('\t', ''))
    return return_list


wb = load_workbook(path_clements)
sheetname = "eBird Clements v2019 Aug 2019"
ws = wb[sheetname]
clements_data = []
clements_species = []
for row in ws.iter_rows(min_row=2, values_only=True):
     data = {'category': row[2], 'english': row[3], 'scientific': row[4]}
     clements_data.append(data)
for item in clements_data:
    if item['category'] == "species":
        bird = {'english': item['english'], 'scientific': item['scientific']}
        clements_species.append(bird)

lst_1 = process_ebird_file(path1)
#lst_2 = process_ebird_file(path2)
#lst_3 = process_ebird_file(path3)
all_ebird_birds = list(set(lst_1))


exotic_birds = []
if path_exotic:
    wb = load_workbook(path_exotic)
    sheetname = "Sheet1"
    ws = wb[sheetname]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            data = {'name': row[1], 'scientific': row[2]}
            exotic_birds.append(data)

targets = []
if path_targets:
    wb = load_workbook(path_targets)
    sheetname = "Sheet1"
    ws = wb[sheetname]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {'name': row[1], 'scientific': row[2]}
        targets.append(data)


sql = "Select BirdName, TaxanomicCode, ScientificName from Birds;"
cursor = conn.cursor()
cursor.execute(sql)
old_birds = cursor.fetchall()

if path_exotic:
    exotic_unmatched_old = []
    for item in exotic_birds:
        flag = False
        for bird in old_birds:
            if bird[2].strip() == item['scientific'].strip():
                flag = True
        if not flag:
            # get english name from clements
            flag2 = False
            for taxon in clements_species:
                if taxon['scientific'].strip() == item['scientific'].strip():
                    flag2 = True
                    english = taxon['english']
                else:
                    english = ''
            diction = {'name': item['name'], 'clements_name': english, 'scientific': item['scientific']}
            exotic_unmatched_old.append(diction)


master_ebird_list = []
for bird in all_ebird_birds:
    flag = False
    code = ''
    scien = ''
    target_value = ''
    add = ''
    # get code from database
    for item in old_birds:
        if bird.strip() == item[0].strip():
            flag = True
            add = ''
            code = item[1]
    if not flag:
        code = ''
        add = 'ADD'
    # get scientific name from clements
    for taxon in clements_species:
        if taxon['english'].strip() == bird.strip():
            scien = taxon['scientific']
    if path_targets:
        # get targets
        flag2 = False
        for target in targets:
            if scien == target['scientific'].strip():
                target_value = "TARGET"
                flag2 = True
        if not flag2:
            target = ''
    else:
        target = ''
    diction = {'code': code, 'name': bird, 'scientific': scien, 'add': add, 'target': target_value}
    master_ebird_list.append(diction)

exotic_not_ebird = []
if path_exotic:
    for exotic in exotic_birds:
        flag = False
        for ebird in master_ebird_list:
            if exotic['scientific'].strip() == ebird['scientific'].strip():
                flag = True
        if not flag:
            # get english name from clements
            flag2 = False
            for taxon in clements_species:
                if taxon['scientific'].strip() == exotic['scientific'].strip():
                    flag2 = True
                    english = taxon['english']
            # Get taxanomic code from old birds if exists
            flag3 = False
            code = ''
            add = ''
            for old in old_birds:
                if exotic['scientific'].strip() == old[2].strip():
                    flag3 = True
                    code = old[1]
            if not flag3:
                code = ''
                add = 'ADD'
            # Get target if exists
            flag4 = False
            for target in targets:
                if exotic['scientific'].strip() == target['scientific'].strip():
                    flag4 = True
                    diction = {'code': code, 'name': english, 'scientific': exotic['scientific'], 'add': add, 'target': 'TARGET'}
                    exotic_not_ebird.append(diction)
            if not flag4:
                diction = {'code': code, 'name': english, 'scientific': exotic['scientific'], 'add': add, 'target': ''}
                exotic_not_ebird.append(diction)

all_island = exotic_not_ebird + master_ebird_list

f = open("palawan_all.csv", "w")
writer = csv.DictWriter(f, fieldnames=["code", "name", "scientific", "add", "target"], lineterminator='\n')
writer.writerows(all_island)
f.close()
